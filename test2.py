import openpyxl

wb = openpyxl.load_workbook('fortnight.xlsx')
sheet = wb['Sheet1']


'''
for row_cells in sheet.iter_rows(min_row=5, max_row=6):
	list1 = []
	for cell in row_cells:
		list1.append(cell.value)
	print(list1)
'''

no = sheet.max_row - 3
print(no)